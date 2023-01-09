import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Font

master_data = xl.load_workbook("master_data.xlsx")
master_sheet = master_data.active
daily_data = xl.load_workbook("daily_data.xlsx")
daily_sheet = daily_data.active


# function to check the number of rows and blank lines
def check_data(sheet_name):
    is_data = True
    row_count = 1
    blank = 0
    while is_data:
        row_count += 1
        info = sheet_name.cell(row, 1).value
        if info == None:
            is_data = False
            blank += 1
            cell = sheet_name.cell(row, 1)
        info = None
    return row, blank, cell


print(check_data(master_sheet))
print(check_data(daily_sheet))

# extracting data from daily sheet and storing it in a list of dictionaries
todays_data = []
for i in range(2, daily_sheet.max_row):
    row_data = {}
    row_data['id'] = daily_sheet.cell(i, 1).value
    row_data['todays_purchase'] = daily_sheet.cell(i, 2).value
    row_data['todays_reward'] = daily_sheet.cell(i, 3).value
    todays_data.append(row_data)

# next task is to append that data(the data we extracted) into master data sheet
# find row using the ID
# go to total purchase cell + today's purchase
# go to total reward balance +today's reward
for i in range(2, master_sheet.max_row):
    ID = master_sheet.cell(i, 1).value
    for row in todays_data:
        if row['id'] == ID:
            todays_purchase = int(row['todays_purchase'])
            todays_reward = int(row['todays_reward'])

            # getting data from master sheet
            total_purchase = master_sheet.cell(i, 6).value
            total_reward = master_sheet.cell(i, 7).value

            # adding values from daily data to total data
            new_total_purchase = todays_purchase + total_purchase
            new_total_reward = todays_reward + total_reward

            master_sheet.cell(i, 6).value = new_total_purchase
            master_sheet.cell(i, 7).value = new_total_reward

master_data.save("master_data.xlsx")

# creating daily report
daily_report = Workbook()
ws = daily_report.active

# get headers
is_Data = True
column_count = 1
header_values = []

while is_Data:
    data = master_sheet.cell(1, column_count).value
    column_count += 1
    if data != None:
        header_values.append(data)
    else:
        is_Data = False

header_style = Font(name="Times New Roman", size=12, bold=True)

# placing headers for the daily report
for i, col_name in enumerate(header_values):
    col_index = i + 1
    ws.cell(1, col_index).value = col_name
    ws.cell(1, col_index).font = header_style

# next task is to append the data from the master data to daily report
# first we grab ids from the daily data
# we had already grabbed them in today's list
IDS = []
for data in todays_data:
    IDS.append(data['id'])
final_data = []
for i in range(2, master_sheet.max_row):
    id = master_sheet.cell(i, 1).value
    if id in IDS:
        lst = []
        for j in range(2, 8):
            lst.append(master_sheet.cell(i, j).value)
        final_data.append(lst)
for data in final_data:
    ws.append(data)

daily_report.save("daily_report.xlsx")
